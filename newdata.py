from openpyxl import load_workbook
from collections import defaultdict


wb1 = load_workbook("Budget 2018.xlsx")

ws_dane = wb1["Nazwy"]

# cell_range = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
#
# for row in ws_dane.iter_rows():
#     print("{}".format([cell.value for cell in row]))

budget_dict = {}
budget_names = []
cols = []
# for row in ws_dane.iter_cols(max_col=1):
#     for cell in row[1:]:
#         budget_dict.setdefault(cell.value, [])
#
# for row in ws_dane.iter_cols(min_col=2, max_col=2):
#     for cell in row[1:]:
#         budget_dict["wydatki"].append(cell.value)
#
# for row in ws_dane.iter_cols(min_col=3, max_col=3):
#     for cell in row[1:]:
#         budget_dict["wydatki"].append(cell.value)

for col in ws_dane.iter_cols():
    for cell in col[1:]:
        if cell.value is not None:
            budget_dict.setdefault(col[0].value.lower(), []).append(cell.value)

# for value in budget_dict["rodzaj"]:
#     budget_names.setdefault(value, {}.setdefault([element for element in budget_dict[value] if element in budget_dict], []))

for k, v in budget_dict.items():
    print(k, v)

for value in budget_dict["rodzaj"]:
    try:
        for name in budget_dict[value]:
            budget_names[value] = budget_dict[name.lower()]
    except KeyError:
        continue
print(budget_names)


